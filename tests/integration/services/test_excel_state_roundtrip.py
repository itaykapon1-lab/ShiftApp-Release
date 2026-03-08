"""Integration tests for exporting session state workbook."""

import pytest
from openpyxl import load_workbook

from data.models import SessionConfigModel
from domain.shift_model import Shift
from domain.worker_model import Worker
from services.excel_service import ExcelService
from domain.time_utils import TimeWindow
from datetime import datetime


pytestmark = [pytest.mark.integration]


def test_constraint_roundtrip_export_then_import(
    db_session, session_id_factory, test_session_id
):
    """Round-trip: constraints survive export_full_state() → import_excel().

    Exercises the full pipeline for avoid_consecutive_shifts and worker_preferences:
      Seed SessionConfigModel (session A) → export_full_state() → ExcelService.import_excel()
      (session B) → assert SessionConfigModel (session B) contains both constraints.
    """
    # Seed session A with constraint types from prior fixes (now resolved)
    constraints = [
        {
            "id": 1,
            "category": "avoid_consecutive_shifts",
            "type": "SOFT",
            "enabled": True,
            "params": {"min_rest_hours": 12, "penalty": -30.0},
        },
        {
            "id": 2,
            "category": "worker_preferences",
            "type": "SOFT",
            "enabled": True,
            "params": {"enabled": True},
        },
    ]
    db_session.add(SessionConfigModel(session_id=test_session_id, constraints=constraints))
    db_session.commit()

    # Export from session A
    service_a = ExcelService(db_session, test_session_id)
    exported_bytes = service_a.export_full_state().read()

    # Import into session B (fresh — no existing constraints to deduplicate against)
    session_b_id = session_id_factory()
    service_b = ExcelService(db_session, session_b_id)
    service_b.import_excel(exported_bytes)

    # Assert session B's constraint DB matches what was exported
    config_b = (
        db_session.query(SessionConfigModel)
        .filter_by(session_id=session_b_id)
        .first()
    )
    assert config_b is not None, "Session B must have a SessionConfigModel after import"
    categories = {c["category"] for c in config_b.constraints}

    assert "avoid_consecutive_shifts" in categories, (
        f"avoid_consecutive_shifts must survive round-trip; got categories: {categories}"
    )
    assert "worker_preferences" in categories, (
        f"worker_preferences must survive round-trip; got categories: {categories}"
    )

    # Verify param fidelity
    avoid = next(
        c for c in config_b.constraints if c["category"] == "avoid_consecutive_shifts"
    )
    assert avoid["params"]["min_rest_hours"] == 12
    assert avoid["params"]["penalty"] == -30.0

    wp = next(
        c for c in config_b.constraints if c["category"] == "worker_preferences"
    )
    assert wp["params"]["enabled"] is True


def test_export_full_state_contains_expected_sheets(db_session, worker_repo, shift_repo, id_factory, test_session_id):
    worker = Worker(name="Alice", worker_id=id_factory("worker"), skills={"Chef": 5})
    shift = Shift(
        name="Morning",
        shift_id=id_factory("shift"),
        time_window=TimeWindow(datetime(2026, 1, 20, 8, 0), datetime(2026, 1, 20, 16, 0)),
    )
    worker_repo.add(worker)
    shift_repo.add(shift)
    db_session.commit()

    service = ExcelService(db_session, test_session_id)
    workbook_bytes = service.export_full_state()
    wb = load_workbook(workbook_bytes)
    assert {"Workers", "Shifts", "Constraints"}.issubset(set(wb.sheetnames))


def test_worker_availability_exports_correct_time_range_not_off(
    db_session, worker_repo, shift_repo, id_factory, test_session_id
):
    """Regression: availability must export as time-range string, never 'OFF'.

    Exercises the full DB round-trip:
      add_availability() -> repo.add() -> DB -> repo.get_all() -> state_exporter -> Excel cell.
    """
    mon_start = datetime(2024, 1, 1, 8, 0)   # Canonical Monday 08:00
    mon_end   = datetime(2024, 1, 1, 16, 0)  # Canonical Monday 16:00
    monday_window = TimeWindow(mon_start, mon_end)

    worker = Worker(name="Alice", worker_id=id_factory("worker"))
    worker.add_availability(mon_start, mon_end)
    worker.add_preference(monday_window, 10)   # score 10 = HIGH -> exports "*"
    worker_repo.add(worker)
    db_session.commit()

    service = ExcelService(db_session, test_session_id)
    wb = load_workbook(service.export_full_state())
    ws = wb["Workers"]

    # Headers: Worker ID | Name | Wage | Min Hours | Max Hours | Skills |
    #          Sunday(7) | Monday(8) | Tuesday(9) | ... | Saturday(13)
    sunday_cell = ws.cell(row=2, column=7).value
    monday_cell = ws.cell(row=2, column=8).value

    assert monday_cell == "08:00-16:00*", (
        f"Monday cell must be '08:00-16:00*' (HIGH pref), got {monday_cell!r}"
    )
    assert sunday_cell == "OFF", (
        f"Sunday cell must be 'OFF' (no availability), got {sunday_cell!r}"
    )


def test_missing_constraints_export_to_constraints_sheet(
    db_session, id_factory, test_session_id
):
    """Regression: avoid_consecutive_shifts and worker_preferences must not be dropped.

    Both categories were silently omitted because they were absent from the
    type_map in _constraint_to_excel_row. This test exercises the full path:
      SessionConfigModel seed -> export_full_state() -> Constraints sheet cell values.
    """
    constraints = [
        {
            "id": 1,
            "category": "avoid_consecutive_shifts",
            "type": "SOFT",
            "enabled": True,
            "params": {"min_rest_hours": 12, "penalty": -30.0},
        },
        {
            "id": 2,
            "category": "worker_preferences",
            "type": "SOFT",
            "enabled": True,
            "params": {"enabled": True},
        },
    ]
    db_session.add(SessionConfigModel(session_id=test_session_id, constraints=constraints))
    db_session.commit()

    service = ExcelService(db_session, test_session_id)
    wb = load_workbook(service.export_full_state())
    ws = wb["Constraints"]

    # Row 2: avoid_consecutive_shifts
    assert ws.cell(row=2, column=1).value == "Avoid Consecutive Shifts", (
        f"Row 2 Type must be 'Avoid Consecutive Shifts', "
        f"got {ws.cell(row=2, column=1).value!r}"
    )
    assert ws.cell(row=2, column=4).value == 12, (
        f"Row 2 Value (min_rest_hours) must be 12, "
        f"got {ws.cell(row=2, column=4).value!r}"
    )
    assert ws.cell(row=2, column=6).value == -30.0, (
        f"Row 2 Penalty must be -30.0, got {ws.cell(row=2, column=6).value!r}"
    )

    # Row 3: worker_preferences
    assert ws.cell(row=3, column=1).value == "Worker Preferences", (
        f"Row 3 Type must be 'Worker Preferences', "
        f"got {ws.cell(row=3, column=1).value!r}"
    )
    assert ws.cell(row=3, column=4).value == "True", (
        f"Row 3 Value (enabled) must be 'True', "
        f"got {ws.cell(row=3, column=4).value!r}"
    )

