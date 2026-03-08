"""
Data Integrity Integration Test Suite
======================================

This module verifies the CRITICAL regression fix where importing Excel data
was wiping out existing manually-entered data. It uses the real production
sample file `Grand_Hotel_Gen_Chaos.xlsx` to prove correctness.

Test Scenarios:
    1. Manual Entry Persistence (Cold Start)
    2. The "Grand Hotel" Non-Destructive Import
    3. Round-Trip Data Resilience
    4. Solver Diagnostics (Logic Check)

Author: QA Automation
Date: 2026-02-12
"""

import os
import io
import uuid
import pytest
from datetime import datetime, timedelta

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, Session

from data.database import Base
from data.models import WorkerModel, ShiftModel, SessionConfigModel
from repositories.sql_repo import SQLWorkerRepository, SQLShiftRepository
from services.excel_service import ExcelService
from domain.worker_model import Worker
from domain.shift_model import Shift
from domain.time_utils import TimeWindow

# ============================================================================
# Constants
# ============================================================================

EXCEL_FILE_PATH = os.path.join(os.path.dirname(__file__), "data", "Grand_Hotel_Gen_Chaos.xlsx")

# Grand Hotel file stats (from direct inspection of the .xlsx)
GRAND_HOTEL_WORKER_COUNT = 77
GRAND_HOTEL_SHIFT_COUNT = 18

# A deterministic session ID for test isolation
TEST_SESSION_ID = "test_data_integrity_session"

# Sample worker names from the Grand Hotel Excel for spot-checking
GRAND_HOTEL_SAMPLE_WORKERS = [
    "Isabella_W001",
    "Olivia_W002",
    "Sophia_W003",
    "William_W004",
    "Emma_W005",
    "Amelia_W013",
    "Elijah_W034",
    "Benjamin_W036",
]

GRAND_HOTEL_SAMPLE_SHIFTS = [
    "Shift_1_Mon",
    "Shift_2_Mon",
    "Shift_3_Sat",
    "Shift_4_Tue",
    "Shift_5_Tue",
]


# ============================================================================
# Fixtures
# ============================================================================

@pytest.fixture(scope="function")
def db_engine():
    """Creates an in-memory SQLite engine per test function."""
    engine = create_engine("sqlite:///:memory:", echo=False)
    Base.metadata.create_all(engine)
    yield engine
    Base.metadata.drop_all(engine)
    engine.dispose()


@pytest.fixture(scope="function")
def db_session(db_engine):
    """Creates a clean SQLAlchemy session per test function."""
    SessionFactory = sessionmaker(bind=db_engine)
    session = SessionFactory()
    yield session
    session.close()


@pytest.fixture(scope="function")
def worker_repo(db_session):
    """Worker repository scoped to the test session."""
    return SQLWorkerRepository(db_session, session_id=TEST_SESSION_ID)


@pytest.fixture(scope="function")
def shift_repo(db_session):
    """Shift repository scoped to the test session."""
    return SQLShiftRepository(db_session, session_id=TEST_SESSION_ID)


@pytest.fixture(scope="function")
def excel_service(db_session):
    """ExcelService instance scoped to the test session."""
    return ExcelService(db_session, session_id=TEST_SESSION_ID)


@pytest.fixture(scope="function")
def grand_hotel_bytes():
    """Reads the real Grand_Hotel_Gen_Chaos.xlsx file into bytes.

    If the file is missing, this fixture will skip the test gracefully.
    """
    if not os.path.isfile(EXCEL_FILE_PATH):
        pytest.skip(
            f"Grand Hotel Excel file not found at: {EXCEL_FILE_PATH}. "
            "Copy Grand_Hotel_Gen_Chaos.xlsx into tests/data/ to enable these tests."
        )
    with open(EXCEL_FILE_PATH, "rb") as f:
        return f.read()


@pytest.fixture(scope="function")
def manual_mike_worker():
    """Creates the canonical 'Manual Mike' domain object for testing."""
    return Worker(
        name="Manual Mike",
        worker_id="MANUAL_MIKE_001",
        wage=30.0,
        min_hours=10,
        max_hours=40,
        skills={"Receptionist": 7, "Security": 3},
    )


@pytest.fixture(scope="function")
def night_watch_shift():
    """Creates the 'Night Watch' shift domain object for testing."""
    base = datetime(2026, 2, 16, 0, 0, 0)  # Monday
    return Shift(
        name="Night Watch",
        time_window=TimeWindow(
            start=base.replace(hour=22),
            end=(base + timedelta(days=1)).replace(hour=6),
        ),
        shift_id="SHIFT_NIGHT_WATCH_001",
    )


@pytest.fixture(scope="function")
def max_hours_constraint():
    """Returns a Max Hours constraint dict in canonical API format."""
    return {
        "id": 999,
        "category": "max_hours_per_week",
        "type": "HARD",
        "name": "Manual Max Hours",
        "description": "Manually configured max hours constraint",
        "params": {"max_hours": 35, "penalty": -50.0},
        "enabled": True,
    }


# ============================================================================
# Helpers
# ============================================================================

def _save_manual_worker(db_session, worker_repo, worker: Worker):
    """Persists a manual worker through the repository layer and commits."""
    worker_repo.add(worker)
    db_session.commit()


def _save_manual_shift(db_session, shift_repo, shift: Shift):
    """Persists a manual shift through the repository layer and commits."""
    shift_repo.add(shift)
    db_session.commit()


def _save_constraint(db_session, session_id: str, constraint: dict):
    """Saves a constraint to the SessionConfigModel."""
    config = db_session.query(SessionConfigModel).filter_by(
        session_id=session_id
    ).first()

    if config:
        existing = config.constraints or []
        existing.append(constraint)
        config.constraints = existing
    else:
        config = SessionConfigModel(
            session_id=session_id,
            constraints=[constraint],
        )
        db_session.add(config)
    db_session.commit()


def _count_workers(db_session, session_id: str) -> int:
    """Counts worker rows for the given session."""
    return db_session.query(WorkerModel).filter(
        WorkerModel.session_id == session_id
    ).count()


def _count_shifts(db_session, session_id: str) -> int:
    """Counts shift rows for the given session."""
    return db_session.query(ShiftModel).filter(
        ShiftModel.session_id == session_id
    ).count()


def _find_worker_by_name(db_session, session_id: str, name: str):
    """Finds a worker by name in the given session."""
    return db_session.query(WorkerModel).filter(
        WorkerModel.session_id == session_id,
        WorkerModel.name == name,
    ).first()


def _find_shift_by_name(db_session, session_id: str, name: str):
    """Finds a shift by name in the given session."""
    return db_session.query(ShiftModel).filter(
        ShiftModel.session_id == session_id,
        ShiftModel.name == name,
    ).first()


# ============================================================================
# TEST 1: Manual Entry Persistence (Cold Start)
# ============================================================================

class TestManualEntryPersistence:
    """Verifies the system handles manual data entry correctly on a fresh DB."""

    def test_worker_persists_with_correct_attributes(
        self, db_session, worker_repo, manual_mike_worker
    ):
        """Manual Mike should be saved and retrievable with all attributes intact."""
        _save_manual_worker(db_session, worker_repo, manual_mike_worker)

        # VERIFY: Worker exists in DB
        db_row = _find_worker_by_name(db_session, TEST_SESSION_ID, "Manual Mike")
        assert db_row is not None, "Manual Mike was not persisted to the database"
        assert db_row.worker_id == "MANUAL_MIKE_001"
        assert db_row.name == "Manual Mike"

        # VERIFY: Attributes JSON blob is well-structured
        attrs = db_row.attributes
        assert attrs is not None, "Worker attributes should not be None"
        assert "skills" in attrs, "Skills should be in attributes"

        # Skills must be persisted (as dict with str keys)
        skills = attrs["skills"]
        assert isinstance(skills, dict), f"Skills should be a dict, got {type(skills)}"
        # Skills get Title-cased by set_skill_level
        assert skills.get("Receptionist", skills.get("receptionist")) is not None
        assert skills.get("Security", skills.get("security")) is not None

    def test_shift_persists_with_correct_window(
        self, db_session, shift_repo, night_watch_shift
    ):
        """Night Watch shift should persist with correct start/end times."""
        _save_manual_shift(db_session, shift_repo, night_watch_shift)

        db_row = _find_shift_by_name(db_session, TEST_SESSION_ID, "Night Watch")
        assert db_row is not None, "Night Watch shift was not persisted"
        assert db_row.shift_id == "SHIFT_NIGHT_WATCH_001"

        # Verify time window
        assert db_row.start_time.hour == 22, "Shift should start at 22:00"
        assert db_row.end_time.hour == 6, "Shift should end at 06:00"
        assert db_row.end_time > db_row.start_time, "End time should be after start time"

    def test_constraint_persists_in_session_config(
        self, db_session, max_hours_constraint
    ):
        """Max Hours constraint should be saved in SessionConfigModel."""
        _save_constraint(db_session, TEST_SESSION_ID, max_hours_constraint)

        config = db_session.query(SessionConfigModel).filter_by(
            session_id=TEST_SESSION_ID
        ).first()
        assert config is not None, "SessionConfigModel should exist"
        assert isinstance(config.constraints, list), "Constraints should be a list"
        assert len(config.constraints) == 1, "Should have exactly 1 constraint"

        saved = config.constraints[0]
        assert saved["category"] == "max_hours_per_week"
        assert saved["params"]["max_hours"] == 35
        assert saved["enabled"] is True

    def test_all_three_entities_coexist(
        self, db_session, worker_repo, shift_repo,
        manual_mike_worker, night_watch_shift, max_hours_constraint,
    ):
        """Worker, Shift, and Constraint should all coexist on the same session."""
        _save_manual_worker(db_session, worker_repo, manual_mike_worker)
        _save_manual_shift(db_session, shift_repo, night_watch_shift)
        _save_constraint(db_session, TEST_SESSION_ID, max_hours_constraint)

        # 3 entities across 3 tables
        assert _count_workers(db_session, TEST_SESSION_ID) == 1
        assert _count_shifts(db_session, TEST_SESSION_ID) == 1

        config = db_session.query(SessionConfigModel).filter_by(
            session_id=TEST_SESSION_ID
        ).first()
        assert config is not None
        assert len(config.constraints) == 1

    def test_repository_domain_roundtrip(
        self, db_session, worker_repo, manual_mike_worker
    ):
        """Workers saved and loaded via repository should return domain objects."""
        _save_manual_worker(db_session, worker_repo, manual_mike_worker)

        # Load via repository (should return domain Worker, not WorkerModel)
        all_workers = worker_repo.get_all()
        assert len(all_workers) == 1, "Should have exactly 1 worker"

        domain_worker = all_workers[0]
        assert isinstance(domain_worker, Worker), (
            f"Repository should return domain Worker, got {type(domain_worker)}"
        )
        assert domain_worker.name == "Manual Mike"
        assert domain_worker.wage == 30.0
        assert domain_worker.max_hours == 40


# ============================================================================
# TEST 2: Non-Destructive Grand Hotel Import
# ============================================================================

class TestNonDestructiveImport:
    """Proves that importing Grand_Hotel_Gen_Chaos.xlsx APPENDS/UPDATES
    but NEVER DELETES existing manual data."""

    def test_manual_mike_survives_import(
        self, db_session, worker_repo, shift_repo, excel_service,
        manual_mike_worker, grand_hotel_bytes,
    ):
        """CRITICAL: Manual Mike MUST still exist after Grand Hotel import.

        This is the primary regression guard. The old code called
        `_clear_session_data()` before import, which would DELETE Manual Mike.
        The fix uses `upsert_by_name` instead.
        """
        # STEP 1: Pre-seed Manual Mike
        _save_manual_worker(db_session, worker_repo, manual_mike_worker)
        pre_import_count = _count_workers(db_session, TEST_SESSION_ID)
        assert pre_import_count == 1, "Pre-condition: Manual Mike must exist"

        # STEP 2: Import Grand Hotel Excel
        result = excel_service.import_excel(grand_hotel_bytes)
        assert result["workers"] > 0, "Import should have found workers"

        # ASSERTION: Manual Mike is STILL HERE
        mike_row = _find_worker_by_name(db_session, TEST_SESSION_ID, "Manual Mike")
        assert mike_row is not None, (
            "REGRESSION DETECTED: Manual Mike was deleted during Excel import! "
            "The import_excel method must use upsert, NOT clear+insert."
        )
        assert mike_row.worker_id == "MANUAL_MIKE_001", (
            "Manual Mike's worker_id was overwritten"
        )

    def test_grand_hotel_workers_imported(
        self, db_session, worker_repo, excel_service, grand_hotel_bytes,
    ):
        """Verifies that sample workers from the Grand Hotel file are present."""
        excel_service.import_excel(grand_hotel_bytes)

        for worker_name in GRAND_HOTEL_SAMPLE_WORKERS:
            row = _find_worker_by_name(db_session, TEST_SESSION_ID, worker_name)
            assert row is not None, (
                f"Grand Hotel worker '{worker_name}' was not imported"
            )

    def test_grand_hotel_shifts_imported(
        self, db_session, shift_repo, excel_service, grand_hotel_bytes,
    ):
        """Verifies that sample shifts from the Grand Hotel file are present."""
        excel_service.import_excel(grand_hotel_bytes)

        for shift_name in GRAND_HOTEL_SAMPLE_SHIFTS:
            row = _find_shift_by_name(db_session, TEST_SESSION_ID, shift_name)
            assert row is not None, (
                f"Grand Hotel shift '{shift_name}' was not imported"
            )

    def test_total_worker_count_is_additive(
        self, db_session, worker_repo, excel_service,
        manual_mike_worker, grand_hotel_bytes,
    ):
        """Total workers = Manual Mike + Grand Hotel workers (no duplicates with name-based upsert)."""
        # Pre-seed Manual Mike
        _save_manual_worker(db_session, worker_repo, manual_mike_worker)

        # Import Grand Hotel
        excel_service.import_excel(grand_hotel_bytes)

        total = _count_workers(db_session, TEST_SESSION_ID)
        # Manual Mike has a unique name (not in Grand Hotel), so total = 1 + 77 = 78
        expected_total = 1 + GRAND_HOTEL_WORKER_COUNT
        assert total == expected_total, (
            f"Expected {expected_total} workers (1 manual + {GRAND_HOTEL_WORKER_COUNT} imported), "
            f"got {total}. Possible: deletion or duplication occurred."
        )

    def test_upsert_updates_existing_worker_by_name(
        self, db_session, worker_repo, excel_service, grand_hotel_bytes,
    ):
        """If a worker with the same name exists, upsert should UPDATE (not duplicate).

        We pre-seed a worker named "Isabella_W001" (same as in Grand Hotel),
        then import. There should be exactly 1 Isabella_W001, not 2.
        """
        # Pre-seed a worker whose name matches a Grand Hotel worker
        existing = Worker(
            name="Isabella_W001",
            worker_id="PRE_EXISTING_ID",
            wage=99.99,
            skills={"Chef": 1},
        )
        _save_manual_worker(db_session, worker_repo, existing)

        # Import Grand Hotel
        excel_service.import_excel(grand_hotel_bytes)

        # Count how many "Isabella_W001" exist
        rows = db_session.query(WorkerModel).filter(
            WorkerModel.session_id == TEST_SESSION_ID,
            WorkerModel.name == "Isabella_W001",
        ).all()

        assert len(rows) == 1, (
            f"Upsert failed: Expected 1 'Isabella_W001', got {len(rows)}. "
            "Worker was duplicated instead of updated."
        )

        # The original worker_id should be preserved (upsert_by_name keeps existing ID)
        assert rows[0].worker_id == "PRE_EXISTING_ID", (
            "Upsert should preserve the existing worker_id, not overwrite it."
        )

    def test_manual_shift_survives_import(
        self, db_session, worker_repo, shift_repo, excel_service,
        night_watch_shift, grand_hotel_bytes,
    ):
        """Night Watch shift should survive the Grand Hotel import."""
        _save_manual_shift(db_session, shift_repo, night_watch_shift)

        excel_service.import_excel(grand_hotel_bytes)

        row = _find_shift_by_name(db_session, TEST_SESSION_ID, "Night Watch")
        assert row is not None, (
            "REGRESSION: Night Watch shift was deleted during Excel import!"
        )

    def test_manual_constraint_survives_import(
        self, db_session, excel_service, max_hours_constraint, grand_hotel_bytes,
    ):
        """Manual constraint should be preserved after Excel import (merge strategy)."""
        _save_constraint(db_session, TEST_SESSION_ID, max_hours_constraint)

        excel_service.import_excel(grand_hotel_bytes)

        config = db_session.query(SessionConfigModel).filter_by(
            session_id=TEST_SESSION_ID
        ).first()
        assert config is not None
        constraints = config.constraints or []

        # Find our manual constraint
        manual_constraints = [
            c for c in constraints
            if c.get("category") == "max_hours_per_week"
            and c.get("params", {}).get("max_hours") == 35
        ]
        assert len(manual_constraints) >= 1, (
            "Manual max_hours_per_week constraint (max=35) was lost during import!"
        )

    def test_double_import_is_idempotent(
        self, db_session, excel_service, grand_hotel_bytes,
    ):
        """Importing the same Excel twice should NOT duplicate data (idempotent)."""
        result1 = excel_service.import_excel(grand_hotel_bytes)
        count_after_first = _count_workers(db_session, TEST_SESSION_ID)

        result2 = excel_service.import_excel(grand_hotel_bytes)
        count_after_second = _count_workers(db_session, TEST_SESSION_ID)

        assert count_after_first == count_after_second, (
            f"Double import created duplicates! "
            f"First: {count_after_first}, Second: {count_after_second}. "
            f"upsert_by_name should prevent this."
        )


# ============================================================================
# TEST 3: Round-Trip Data Resilience
# ============================================================================

class TestRoundTripResilience:
    """Ensures we can export the hybrid state and re-import it faithfully."""

    def test_export_produces_valid_excel_bytes(
        self, db_session, worker_repo, shift_repo, excel_service,
        manual_mike_worker, night_watch_shift, grand_hotel_bytes,
    ):
        """export_full_state() should return a BytesIO with valid Excel content."""
        # Build hybrid state
        _save_manual_worker(db_session, worker_repo, manual_mike_worker)
        _save_manual_shift(db_session, shift_repo, night_watch_shift)
        excel_service.import_excel(grand_hotel_bytes)

        # Export
        exported = excel_service.export_full_state()
        assert isinstance(exported, io.BytesIO), "Export should return BytesIO"

        exported_bytes = exported.getvalue()
        assert len(exported_bytes) > 0, "Exported file should not be empty"

        # Verify it's a valid .xlsx (ZIP magic bytes: PK)
        assert exported_bytes[:2] == b"PK", (
            "Exported file does not have XLSX/ZIP signature"
        )

    def test_export_contains_all_workers_and_shifts(
        self, db_session, worker_repo, shift_repo, excel_service,
        manual_mike_worker, night_watch_shift, grand_hotel_bytes,
    ):
        """The exported Excel should contain rows for all workers and shifts."""
        import openpyxl

        # Build hybrid state
        _save_manual_worker(db_session, worker_repo, manual_mike_worker)
        _save_manual_shift(db_session, shift_repo, night_watch_shift)
        excel_service.import_excel(grand_hotel_bytes)

        # Get counts before export
        total_workers = _count_workers(db_session, TEST_SESSION_ID)
        total_shifts = _count_shifts(db_session, TEST_SESSION_ID)

        # Export & inspect
        exported = excel_service.export_full_state()
        wb = openpyxl.load_workbook(exported)

        assert "Workers" in wb.sheetnames, "Exported file must have 'Workers' sheet"
        assert "Shifts" in wb.sheetnames, "Exported file must have 'Shifts' sheet"

        ws_workers = wb["Workers"]
        ws_shifts = wb["Shifts"]

        # -1 for header row
        exported_worker_count = ws_workers.max_row - 1
        exported_shift_count = ws_shifts.max_row - 1

        assert exported_worker_count == total_workers, (
            f"Exported worker count ({exported_worker_count}) != "
            f"DB worker count ({total_workers})"
        )
        assert exported_shift_count == total_shifts, (
            f"Exported shift count ({exported_shift_count}) != "
            f"DB shift count ({total_shifts})"
        )

    def test_reimport_after_wipe_restores_counts(
        self, db_session, db_engine, worker_repo, shift_repo, excel_service,
        manual_mike_worker, night_watch_shift, grand_hotel_bytes,
    ):
        """Full round-trip: Seed → Import → Export → Wipe → Re-import → Verify counts."""
        # STEP 1: Build hybrid state (Manual + Grand Hotel)
        _save_manual_worker(db_session, worker_repo, manual_mike_worker)
        _save_manual_shift(db_session, shift_repo, night_watch_shift)
        excel_service.import_excel(grand_hotel_bytes)

        expected_workers = _count_workers(db_session, TEST_SESSION_ID)
        expected_shifts = _count_shifts(db_session, TEST_SESSION_ID)
        assert expected_workers > 1, "Should have more than 1 worker"
        assert expected_shifts > 1, "Should have more than 1 shift"

        # STEP 2: Export full state
        exported = excel_service.export_full_state()
        exported_bytes = exported.getvalue()

        # STEP 3: Simulate DB wipe (delete all data for this session)
        db_session.query(WorkerModel).filter(
            WorkerModel.session_id == TEST_SESSION_ID
        ).delete()
        db_session.query(ShiftModel).filter(
            ShiftModel.session_id == TEST_SESSION_ID
        ).delete()
        db_session.commit()

        # Verify wipe
        assert _count_workers(db_session, TEST_SESSION_ID) == 0, "DB should be empty"
        assert _count_shifts(db_session, TEST_SESSION_ID) == 0, "DB should be empty"

        # STEP 4: Re-import the exported file
        restored_service = ExcelService(db_session, session_id=TEST_SESSION_ID)
        result = restored_service.import_excel(exported_bytes)

        # STEP 5: Verify counts match the hybrid state
        restored_workers = _count_workers(db_session, TEST_SESSION_ID)
        restored_shifts = _count_shifts(db_session, TEST_SESSION_ID)

        assert restored_workers == expected_workers, (
            f"Round-trip worker loss! Expected {expected_workers}, "
            f"got {restored_workers}. Export format may not match import parser."
        )
        assert restored_shifts == expected_shifts, (
            f"Round-trip shift loss! Expected {expected_shifts}, "
            f"got {restored_shifts}. Export format may not match import parser."
        )

    def test_manual_mike_name_survives_roundtrip(
        self, db_session, db_engine, worker_repo, shift_repo, excel_service,
        manual_mike_worker, grand_hotel_bytes,
    ):
        """Manual Mike should be findable by name after a full round-trip."""
        # Build → Export → Wipe → Re-import
        _save_manual_worker(db_session, worker_repo, manual_mike_worker)
        excel_service.import_excel(grand_hotel_bytes)
        exported = excel_service.export_full_state()
        exported_bytes = exported.getvalue()

        # Wipe
        db_session.query(WorkerModel).filter(
            WorkerModel.session_id == TEST_SESSION_ID
        ).delete()
        db_session.commit()

        # Re-import
        restored_service = ExcelService(db_session, session_id=TEST_SESSION_ID)
        restored_service.import_excel(exported_bytes)

        # Mike should exist
        mike = _find_worker_by_name(db_session, TEST_SESSION_ID, "Manual Mike")
        assert mike is not None, (
            "Manual Mike was lost during round-trip (export → wipe → reimport)"
        )


# ============================================================================
# TEST 4: Solver Diagnostics (Logic Check)
# ============================================================================

class TestSolverDiagnostics:
    """Verifies the solver correctly identifies infeasible schedules."""

    def test_conflicting_mutual_exclusion_causes_infeasibility(
        self, db_session, worker_repo, shift_repo, excel_service, grand_hotel_bytes,
    ):
        """When the ONLY two eligible workers for a shift are mutually excluded,
        the solver should report 'Infeasible' and provide a diagnosis.

        Strategy:
            1. Create a minimal scenario with exactly 2 workers.
            2. Create a shift requiring exactly 2 workers with those skills.
            3. Ban them from working together (Mutual Exclusion HARD).
            4. This makes it structurally impossible to staff the shift.
        """
        from services.session_adapter import SessionDataManagerAdapter
        from solver.solver_engine import ShiftSolver
        from solver.constraints.registry import ConstraintRegistry
        from solver.constraints.dynamic import MutualExclusionConstraint
        from solver.constraints.base import ConstraintType
        from domain.task_model import Task, TaskOption

        # Worker A: has exactly the skills needed
        worker_a = Worker(
            name="Alice Conflict",
            worker_id="CONFLICT_A",
            skills={"TestSkill": 5},
        )
        # Worker B: has exactly the skills needed
        worker_b = Worker(
            name="Bob Conflict",
            worker_id="CONFLICT_B",
            skills={"TestSkill": 5},
        )

        # Both are available at the same time
        base_dt = datetime(2026, 2, 16, 8, 0, 0)  # Monday 8:00
        end_dt = datetime(2026, 2, 16, 16, 0, 0)   # Monday 16:00
        worker_a.add_availability(base_dt, end_dt)
        worker_b.add_availability(base_dt, end_dt)

        # Shift requiring 2 workers with TestSkill
        time_window = TimeWindow(base_dt, end_dt)
        shift = Shift(
            name="Impossible Shift",
            time_window=time_window,
            shift_id="SHIFT_IMPOSSIBLE",
        )

        # Task: Requires 2 workers with TestSkill:5
        task = Task(name="Impossible Task")
        option = TaskOption()
        option.add_requirement(count=2, required_skills={"TestSkill": 5})
        task.add_option(option)
        shift.add_task(task)

        # Build the data adapter
        adapter = SessionDataManagerAdapter(
            workers=[worker_a, worker_b],
            shifts=[shift],
        )

        # Build constraint registry with mutual exclusion
        registry = ConstraintRegistry()
        registry.add_core_constraints()
        registry.register(
            MutualExclusionConstraint(
                worker_a_id="CONFLICT_A",
                worker_b_id="CONFLICT_B",
                strictness=ConstraintType.HARD,
            )
        )

        # Run solver
        solver = ShiftSolver(adapter, constraint_registry=registry)
        result = solver.solve()

        # ASSERTION: Should be infeasible
        assert result["status"] == "Infeasible", (
            f"Expected 'Infeasible' but got '{result['status']}'. "
            "The mutual exclusion constraint should make it impossible to staff "
            "a shift requiring 2 workers when they cannot work together."
        )

    def test_diagnosis_identifies_culprit_constraint(
        self, db_session,
    ):
        """diagnose_infeasibility() should return a message identifying the
        specific constraint that caused the failure.
        """
        from services.session_adapter import SessionDataManagerAdapter
        from solver.solver_engine import ShiftSolver
        from solver.constraints.registry import ConstraintRegistry
        from solver.constraints.dynamic import MutualExclusionConstraint
        from solver.constraints.base import ConstraintType
        from domain.task_model import Task, TaskOption

        # Same conflicting setup as above
        worker_a = Worker(
            name="Diag Alice",
            worker_id="DIAG_A",
            skills={"Cooking": 5},
        )
        worker_b = Worker(
            name="Diag Bob",
            worker_id="DIAG_B",
            skills={"Cooking": 5},
        )

        base_dt = datetime(2026, 2, 16, 8, 0, 0)
        end_dt = datetime(2026, 2, 16, 16, 0, 0)
        worker_a.add_availability(base_dt, end_dt)
        worker_b.add_availability(base_dt, end_dt)

        tw = TimeWindow(base_dt, end_dt)
        shift = Shift(name="Diag Shift", time_window=tw, shift_id="SHIFT_DIAG")

        task = Task(name="Diagnostic Task")
        option = TaskOption()
        option.add_requirement(count=2, required_skills={"Cooking": 5})
        task.add_option(option)
        shift.add_task(task)

        adapter = SessionDataManagerAdapter(
            workers=[worker_a, worker_b],
            shifts=[shift],
        )

        registry = ConstraintRegistry()
        registry.add_core_constraints()
        registry.register(
            MutualExclusionConstraint(
                worker_a_id="DIAG_A",
                worker_b_id="DIAG_B",
                strictness=ConstraintType.HARD,
            )
        )

        solver = ShiftSolver(adapter, constraint_registry=registry)
        result = solver.solve()
        assert result["status"] == "Infeasible"

        # Now run diagnosis
        diagnosis = solver.diagnose_infeasibility()
        assert isinstance(diagnosis, str), "Diagnosis should return a string"
        assert len(diagnosis) > 0, "Diagnosis message should not be empty"

        # The diagnosis should mention the ban constraint or its effects
        diagnosis_lower = diagnosis.lower()
        assert any(keyword in diagnosis_lower for keyword in [
            "ban", "mutual", "exclusion", "conflict", "constraint",
            "infeasib", "failure", "impossible",
        ]), (
            f"Diagnosis message doesn't identify the constraint. Got: {diagnosis}"
        )

    def test_feasible_scenario_without_conflicts(self, db_session):
        """Sanity check: A valid scenario without conflicts should be solvable."""
        from services.session_adapter import SessionDataManagerAdapter
        from solver.solver_engine import ShiftSolver
        from domain.task_model import Task, TaskOption

        # Two workers, each can handle the job alone
        worker_a = Worker(
            name="Happy Alice",
            worker_id="HAPPY_A",
            skills={"Cleaning": 5},
        )
        worker_b = Worker(
            name="Happy Bob",
            worker_id="HAPPY_B",
            skills={"Cleaning": 5},
        )

        base_dt = datetime(2026, 2, 16, 8, 0, 0)
        end_dt = datetime(2026, 2, 16, 16, 0, 0)
        worker_a.add_availability(base_dt, end_dt)
        worker_b.add_availability(base_dt, end_dt)

        tw = TimeWindow(base_dt, end_dt)
        shift = Shift(name="Easy Shift", time_window=tw, shift_id="SHIFT_EASY")

        # Only need 1 worker
        task = Task(name="Easy Task")
        option = TaskOption()
        option.add_requirement(count=1, required_skills={"Cleaning": 5})
        task.add_option(option)
        shift.add_task(task)

        adapter = SessionDataManagerAdapter(
            workers=[worker_a, worker_b],
            shifts=[shift],
        )

        solver = ShiftSolver(adapter)
        result = solver.solve()

        assert result["status"] in ("Optimal", "Feasible"), (
            f"Expected solvable but got '{result['status']}'"
        )
        assert len(result["assignments"]) >= 1, (
            "Should have at least 1 assignment"
        )

    def test_solver_with_grand_hotel_data(
        self, db_session, worker_repo, shift_repo, excel_service, grand_hotel_bytes,
    ):
        """The Grand Hotel dataset should be processable by the solver.

        We don't assert Optimal here (the Chaos dataset may be intentionally
        complex), but the solver should at least produce a result without crashing.
        """
        from services.session_adapter import SessionDataManagerAdapter
        from solver.solver_engine import ShiftSolver
        from solver.constraints.registry import ConstraintRegistry

        # Import Grand Hotel data
        excel_service.import_excel(grand_hotel_bytes)

        # Load domain objects through repos
        workers = worker_repo.get_all()
        shifts = shift_repo.get_all()

        assert len(workers) > 0, "Should have workers after import"
        assert len(shifts) > 0, "Should have shifts after import"

        # Build adapter and solver
        adapter = SessionDataManagerAdapter(workers=workers, shifts=shifts)
        registry = ConstraintRegistry()
        registry.add_core_constraints()

        solver = ShiftSolver(adapter, constraint_registry=registry)
        result = solver.solve()

        # The result should be a valid dict with expected keys
        assert "status" in result
        assert "assignments" in result
        assert result["status"] in ("Optimal", "Feasible", "Infeasible"), (
            f"Unexpected solver status: {result['status']}"
        )

        # If feasible, we should have assignments
        if result["status"] in ("Optimal", "Feasible"):
            assert len(result["assignments"]) > 0, (
                "Solver returned Feasible/Optimal but no assignments"
            )
