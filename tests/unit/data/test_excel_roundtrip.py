"""Round-trip tests for prioritized task options through Excel export/import."""

import datetime as dt
import io

import pytest
from openpyxl import load_workbook

from domain.shift_model import Shift
from domain.task_model import Task, TaskOption
from repositories.sql_shift_repo import SQLShiftRepository
from services.excel_service import ExcelService
from domain.time_utils import TimeWindow


pytestmark = [pytest.mark.unit]


def _build_priority_shift() -> Shift:
    start = dt.datetime(2026, 2, 16, 8, 0, 0)
    end = dt.datetime(2026, 2, 16, 16, 0, 0)

    shift = Shift(
        name="Priority Roundtrip Shift",
        shift_id="S_PRIORITY_ROUNDTRIP",
        time_window=TimeWindow(start, end),
    )
    task = Task(name="Kitchen", task_id="T_PRIORITY")

    option_1 = TaskOption(priority=1)
    option_1.add_requirement(count=1, required_skills={"Chef": 5})
    option_2 = TaskOption(priority=2)
    option_2.add_requirement(count=1, required_skills={"Cook": 3})

    task.add_option(option_1)
    task.add_option(option_2)
    shift.add_task(task)
    return shift


def test_priority_roundtrip_export_then_import_preserves_option_priorities(
    db_session, session_id_factory
):
    source_session_id = session_id_factory()
    source_shift_repo = SQLShiftRepository(db_session, source_session_id)
    source_shift_repo.add(_build_priority_shift())
    db_session.commit()

    source_service = ExcelService(db_session, source_session_id)
    exported_bytes = source_service.export_full_state().read()

    workbook = load_workbook(io.BytesIO(exported_bytes))
    shifts_sheet = workbook["Shifts"]
    tasks_cell = shifts_sheet.cell(row=2, column=5).value
    assert "#1:" in tasks_cell
    assert "#2:" in tasks_cell

    target_session_id = session_id_factory()
    target_service = ExcelService(db_session, target_session_id)
    target_service.import_excel(exported_bytes)

    target_shift_repo = SQLShiftRepository(db_session, target_session_id)
    imported_shift = next(
        shift for shift in target_shift_repo.get_all() if shift.name == "Priority Roundtrip Shift"
    )

    assert len(imported_shift.tasks) == 1
    assert len(imported_shift.tasks[0].options) == 2
    assert [option.priority for option in imported_shift.tasks[0].options] == [1, 2]
