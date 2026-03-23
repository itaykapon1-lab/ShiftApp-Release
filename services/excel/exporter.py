"""
Excel Exporter — Generates Excel file with schedule results.

Extracted from services/excel_service.py.
All logic, variables, magic numbers, and comments are preserved exactly.
"""

import io
import logging

import pandas as pd
from sqlalchemy.orm import Session

# Repositories
from repositories.sql_repo import SQLWorkerRepository, SQLShiftRepository

# Configure logger
logger = logging.getLogger(__name__)


class ExcelExporter:
    def __init__(self, db: Session, session_id: str,
                 worker_repo: SQLWorkerRepository,
                 shift_repo: SQLShiftRepository):
        self.db = db
        self.session_id = session_id
        self.worker_repo = worker_repo
        self.shift_repo = shift_repo

    def export_excel(self) -> io.BytesIO:
        """
        Generates Excel file with exact flat table format for schedule.
        (Logic unchanged from original upload, kept for completeness)
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Load all workers and shifts to build ID→Name lookups.
        # The solver stores assignments by ID; the Excel output needs human names.
        workers = self.worker_repo.get_all()
        shifts = self.shift_repo.get_all()

        worker_map = {w.worker_id: w.name for w in workers}   # worker_id → display name
        shift_map = {s.shift_id: s.name for s in shifts}      # shift_id → display name

        # Fetch the most recent solver results from the DB-backed job store.
        # This import is deferred to avoid circular imports (solver_service → excel_service).
        from services.solver_service import SolverService

        assignments_data = []
        latest_job = SolverService.get_latest_job_for_session(self.session_id)

        if latest_job:
            raw_assignments = latest_job.get("assignments", [])

            # Transform each solver assignment dict into a flat row for the
            # Schedule sheet.  IDs are resolved to names via the lookup dicts.
            for assign in raw_assignments:
                worker_id = assign.get('worker_id', 'Unknown')
                shift_id = assign.get('shift_id', 'Unknown')
                worker_name = worker_map.get(worker_id, worker_id)  # Fallback to ID if unknown
                shift_name = shift_map.get(shift_id, shift_id)
                # Parse the ISO datetime string into separate Date and Time columns.
                time_str = assign.get('time', '')
                date_part = time_str[:10] if len(time_str) >= 10 else 'N/A'   # "2024-01-01"
                time_part = time_str[11:16] if len(time_str) >= 16 else 'N/A'  # "08:00"

                assignments_data.append({
                    'Date': date_part,
                    'Time': time_part,
                    'Worker': worker_name,
                    'Shift': shift_name,
                    'Score': assign.get('score', 0)  # Soft-constraint preference score
                })

        # Write the Schedule sheet to an in-memory Excel workbook.
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Professional styling — dark blue header, white text, thin borders.
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            center_align = Alignment(horizontal='center', vertical='center')

            if assignments_data:
                # Build a DataFrame sorted chronologically, then by worker name.
                df_schedule = pd.DataFrame(assignments_data, columns=['Date', 'Time', 'Worker', 'Shift', 'Score'])
                df_schedule = df_schedule.sort_values(by=['Date', 'Time', 'Worker'])
                df_schedule.to_excel(writer, sheet_name='Schedule', index=False)

                # Apply header styling to the first row.
                ws = writer.sheets['Schedule']
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = border_thin

                # Apply borders to all data rows.
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.border = border_thin
                        cell.alignment = Alignment(vertical='center')

                # Auto-fit column widths based on cell content (capped at 40 chars).
                for col_idx in range(1, ws.max_column + 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = 10
                    for cell in ws[column_letter]:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
            else:
                # No solver results — write a single informational row.
                pd.DataFrame({'Info': ['No solver results found. Please run the solver first.']}).to_excel(writer, sheet_name='Schedule', index=False)

        # Reset the buffer position so the caller can read from the beginning.
        output.seek(0)
        return output
