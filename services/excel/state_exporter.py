"""
State Exporter — Exports complete session state as round-trip compatible Excel.

Extracted from services/excel_service.py.
All logic, variables, magic numbers, and comments are preserved exactly.
"""

import io
import logging

from sqlalchemy.orm import Session

# Database Models
from data.models import SessionConfigModel

# Repositories
from repositories.sql_repo import SQLWorkerRepository, SQLShiftRepository

# Configure logger
logger = logging.getLogger(__name__)


class StateExporter:
    def __init__(self, db: Session, session_id: str,
                 worker_repo: SQLWorkerRepository,
                 shift_repo: SQLShiftRepository):
        self.db = db
        self.session_id = session_id
        self.worker_repo = worker_repo
        self.shift_repo = shift_repo

    def export_full_state(self) -> io.BytesIO:
        """Export complete session state as round-trip compatible Excel.

        Creates an Excel file with Workers, Shifts, and Constraints sheets
        in the same format as the import expects, enabling round-trip
        (export -> modify -> re-import) workflows.

        Returns:
            io.BytesIO: Excel file buffer
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Create a new workbook with three sheets that mirror the import format.
        # This enables round-trip workflows: export → edit in Excel → re-import.
        wb = Workbook()

        # Professional styling — matches the schedule export aesthetics.
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Sheet 1: Workers — same column layout the importer expects.
        ws_workers = wb.active
        ws_workers.title = "Workers"
        self._write_workers_sheet(ws_workers, header_fill, header_font, border_thin)

        # Sheet 2: Shifts — Day, Shift Name, Start Time, End Time, Tasks.
        ws_shifts = wb.create_sheet("Shifts")
        self._write_shifts_sheet(ws_shifts, header_fill, header_font, border_thin)

        # Sheet 3: Constraints — Type, Subject, Target, Value, Strictness, Penalty.
        ws_constraints = wb.create_sheet("Constraints")
        self._write_constraints_sheet(ws_constraints, header_fill, header_font, border_thin)

        # Auto-adjust column widths for all sheets
        for ws in [ws_workers, ws_shifts, ws_constraints]:
            for col_idx in range(1, ws.max_column + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 10
                for cell in ws[column_letter]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _write_workers_sheet(self, ws, header_fill, header_font, border_thin):
        """Write workers in import-compatible format."""
        headers = ["Worker ID", "Name", "Wage", "Min Hours", "Max Hours",
                   "Skills", "Sunday", "Monday", "Tuesday", "Wednesday",
                   "Thursday", "Friday", "Saturday"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border_thin

        workers = self.worker_repo.get_all()
        for w in workers:
            # Format skills as comma-separated "Skill:Level" pairs.
            # Example: "Cook:5,Waiter:3" — matches the import parser's expectation.
            skills_str = ",".join(f"{s}:{l}" for s, l in w.skills.items()) if w.skills else ""

            # Base row: static worker attributes.
            row = [
                w.worker_id, w.name, w.wage, w.min_hours, w.max_hours, skills_str
            ]

            # Convert the worker's availability TimeWindows and preferences back
            # into the dict format used by the Excel columns (one column per day).
            raw_avail = self.worker_repo._convert_availability_to_dict_format(
                w.availability, w.preferences
            )
            # Day abbreviations match the keys used in raw_avail; full names are
            # for the column headers (Sunday through Saturday).
            days = ["SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT"]
            day_full = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

            for day_abbr, day_name in zip(days, day_full):
                day_data = raw_avail.get(day_abbr, None)
                if not day_data:
                    row.append("OFF")  # No availability = day off
                elif isinstance(day_data, str):
                    row.append(day_data)  # Already formatted (e.g., "08:00-16:00")
                elif isinstance(day_data, dict):
                    time_range = day_data.get("timeRange", "08:00-16:00")
                    pref = day_data.get("preference", "NEUTRAL")
                    # Append preference markers that the importer recognises:
                    if pref == "HIGH":
                        time_range += "*"  # * = worker prefers this shift (+bonus)
                    elif pref == "LOW":
                        time_range += "!"  # ! = worker wants to avoid this shift (-penalty)
                    row.append(time_range)
                else:
                    row.append("OFF")

            ws.append(row)

            # Style data rows
            for cell in ws[ws.max_row]:
                cell.border = border_thin

    def _write_shifts_sheet(self, ws, header_fill, header_font, border_thin):
        """Write shifts in import-compatible format."""
        headers = ["Day", "Shift Name", "Start Time", "End Time", "Tasks"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border_thin

        shifts = self.shift_repo.get_all()
        for s in shifts:
            # Derive the day-of-week name from the canonical epoch datetime.
            # Python's weekday(): Mon=0..Sun=6, so this maps correctly.
            day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
            day = day_names[s.time_window.start.weekday()]

            # Format times.
            # Same-day shifts: straightforward HH:MM format.
            # Overnight shifts (end date > start date) use the +24h convention:
            # e.g. a shift starting Monday 22:00 and ending Tuesday 06:00 is
            # written as Start="22:00", End="30:00".  This is lossless and
            # unambiguous — the validator and parser both accept this notation.
            start_str = s.time_window.start.strftime("%H:%M")
            if s.time_window.end.date() > s.time_window.start.date():
                # Overnight: add 24 to the end hour to produce +24h notation.
                overnight_hours = s.time_window.end.hour + 24
                end_str = f"{overnight_hours:02d}:{s.time_window.end.strftime('%M')}"
            else:
                end_str = s.time_window.end.strftime("%H:%M")

            # Serialize the shift's task definitions back to the grammar string
            # format: "#1: [Skill:Level] x Count | #2: [Skill2:Level] x Count2"
            tasks_str = self._serialize_tasks_to_string(s.tasks) if s.tasks else ""

            ws.append([day, s.name, start_str, end_str, tasks_str])

            # Style data rows
            for cell in ws[ws.max_row]:
                cell.border = border_thin

    def _serialize_tasks_to_string(self, tasks) -> str:
        """Convert tasks structure to Excel string format.

        Uses canonical ``#X:`` prefix syntax for every option so exported files
        re-import cleanly without legacy-syntax warnings.
        Multiple tasks within a shift are joined by `` | ``.

        Args:
            tasks: List of Task domain objects.

        Returns:
            str: Task string in import-compatible format.
        """
        if not tasks:
            return ""

        task_parts = []
        for task in tasks:
            option_parts = []
            for option_index, option in enumerate(task.options, start=1):
                # Build the requirement brackets for this option.
                # Each requirement: "[Skill:Level, Skill2:Level] x Count"
                req_parts = []
                for req in option.requirements:
                    skills_str = ", ".join(
                        f"{name}:{level}" for name, level in req.required_skills.items()
                    ) if req.required_skills else ""
                    req_parts.append(f"[{skills_str}] x {req.count}")
                # Join multiple requirements within one option with " + " (AND).
                body = " + ".join(req_parts) if req_parts else "[] x 1"

                # Use the option's stored priority, falling back to its positional
                # index if the priority is missing or invalid.
                priority = getattr(option, 'priority', None)
                if not isinstance(priority, int) or priority < 1:
                    priority = option_index
                # Prefix with "#N:" so the parser can reconstruct option ordering.
                option_parts.append(f"#{priority}: {body}")

            if option_parts:
                # Options within a task are space-separated (the parser splits on #N:).
                task_parts.append(" ".join(option_parts))

        # Tasks within a shift are pipe-separated.
        return " | ".join(task_parts)

    def _write_constraints_sheet(self, ws, header_fill, header_font, border_thin):
        """Write constraints in import-compatible format."""
        headers = ["Type", "Subject", "Target", "Value", "Strictness", "Penalty"]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border_thin

        # Load the session's persisted constraint config from the DB.
        # Each constraint is a JSON dict with {category, type, params, ...}.
        config = self.db.query(SessionConfigModel).filter_by(
            session_id=self.session_id
        ).first()

        if not config or not config.constraints:
            return  # No constraints to export

        constraints = config.constraints if isinstance(config.constraints, list) else []

        # Convert each canonical constraint back to the flat Excel row format.
        for c in constraints:
            row = self._constraint_to_excel_row(c)
            if row:
                ws.append(row)
                # Style data rows
                for cell in ws[ws.max_row]:
                    cell.border = border_thin

    def _constraint_to_excel_row(self, constraint: dict) -> list:
        """Convert canonical JSON constraint to Excel row format.

        Args:
            constraint: Constraint dict in canonical format

        Returns:
            list: [Type, Subject, Target, Value, Strictness, Penalty] or None
        """
        category = constraint.get("category")   # Canonical key, e.g., "mutual_exclusion"
        params = constraint.get("params", {})    # Constraint-specific parameters
        strictness_raw = str(constraint.get("type", "HARD")).strip().upper()
        strictness = strictness_raw if strictness_raw in {"HARD", "SOFT"} else "HARD"

        # Reverse mapping: canonical category → user-friendly Excel type name.
        # This is the inverse of the importer's raw_type_lower matching.
        type_map = {
            "mutual_exclusion":         "Mutual Exclusion",
            "colocation":               "Co-Location",
            "max_hours_per_week":       "Max Hours",
            "avoid_consecutive_shifts": "Avoid Consecutive Shifts",
            "worker_preferences":       "Worker Preferences",
            "task_option_priority":     "Task Option Priority",
        }

        excel_type = type_map.get(category)
        if not excel_type:
            return None  # Unrecognised category — skip silently

        # Each constraint category maps its params to specific Excel columns:
        # [Type, Subject, Target, Value, Strictness, Penalty]

        if category == "mutual_exclusion":
            # Subject = Worker A, Target = Worker B (the banned pair)
            return [
                excel_type,
                params.get("worker_a_id", ""),
                params.get("worker_b_id", ""),
                "",  # Value not used for ban constraints
                strictness,
                params.get("penalty", -100)
            ]
        elif category == "colocation":
            # Subject = Worker A, Target = Worker B (the required pair)
            return [
                excel_type,
                params.get("worker_a_id", ""),
                params.get("worker_b_id", ""),
                "",  # Value not used for pair constraints
                strictness,
                params.get("penalty", -100)
            ]
        elif category == "max_hours_per_week":
            # Value = the hour cap (e.g., 40)
            return [
                excel_type,
                "",  # No subject (applies to all workers)
                "",  # No target
                params.get("max_hours", 40),
                strictness,
                params.get("penalty", -50)
            ]
        elif category == "avoid_consecutive_shifts":
            # Value = minimum rest hours between shifts
            return [
                excel_type,
                "",
                "",
                params.get("min_rest_hours", 12),
                strictness,
                params.get("penalty", -30.0),
            ]
        elif category == "worker_preferences":
            # Subject = preference reward weight, Value = enabled toggle
            return [
                excel_type,
                params.get("preference_reward", 10),
                "",
                str(params.get("enabled", True)),
                strictness,
                params.get("preference_penalty", -100),
            ]
        elif category == "task_option_priority":
            # Always SOFT — penalises non-preferred task options
            return [
                excel_type,
                "",
                "",
                "",
                "SOFT",
                params.get("base_penalty", -20.0),
            ]

        return None
